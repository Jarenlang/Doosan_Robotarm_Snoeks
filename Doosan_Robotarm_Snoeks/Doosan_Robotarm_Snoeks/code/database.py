import openpyxl
from pathlib import Path

# Pad naar het Excel‑bestand
BASEDIR = Path(__file__).resolve().parent
DATADIR = BASEDIR / ".." / "data"
WORKORDERS_PATH = DATADIR / "Workorders.xlsx"

# Welke sheet en kolommen
SHEET_INDEX = 1            # "pagina twee" = tweede werkblad (index 1)
COL_WORKORDER = "A"        # voorbeeld: kolom met WORKORDERBASEID
COL_PARTID = "E"           # kolom PARTID
COL_TRACEID = "J"          # kolom TRACEID
COL_DESCRIPTION = "H"


class PartNumberError(Exception):
    pass


def _load_sheet():
    if not WORKORDERS_PATH.exists():
        raise FileNotFoundError(f"Workorders.xlsx niet gevonden op {WORKORDERS_PATH}")
    wb = openpyxl.load_workbook(WORKORDERS_PATH)
    ws = wb.worksheets[SHEET_INDEX]
    return wb, ws


def _find_rows_for_workorder(ws, workorder_id: str):

    rows = []
    for row in range(2, ws.max_row + 1):  # aannemen rij 1 = header
        cell_value = str(ws[f"{COL_WORKORDER}{row}"].value or "").strip()
        if cell_value == workorder_id:
            rows.append(row)
    return rows


def _get_partids_for_rows(ws, rows):
    partids = set()
    for r in rows:
        val = ws[f"{COL_PARTID}{r}"].value
        if val is not None:
            partids.add(str(val).strip())
    return partids

def validate_workorder_exists(workorder_id: str) -> None:
    wb, ws = _load_sheet()
    rows = _find_rows_for_workorder(ws, workorder_id)
    if not rows:
        raise ValueError(
            f"Workorder '{workorder_id}' komt niet voor op pagina twee van Workorders.xlsx."
        )


def validate_scanned_parts(workorder_id: str,
                           scanned_frame_part: str | None,
                           scanned_belt_part: str | None,
                           scanned_buckle_part: str | None):
    wb, ws = _load_sheet()
    rows = _find_rows_for_workorder(ws, workorder_id)
    if not rows:
        raise PartNumberError(f"Geen rijen gevonden voor workorder '{workorder_id}' op pagina twee.")

    db_parts = _get_partids_for_rows(ws, rows)

    # Controle frame
    if scanned_frame_part and scanned_frame_part not in db_parts:
        raise PartNumberError(
            f"Frame-partnummer '{scanned_frame_part}' komt niet overeen met database voor workorder {workorder_id}."
        )

    if scanned_belt_part and scanned_belt_part not in db_parts:
        raise PartNumberError(
            f"Seatbelt-partnummer '{scanned_belt_part}' komt niet overeen met database voor workorder {workorder_id}."
        )

    if scanned_buckle_part and scanned_buckle_part not in db_parts:
        raise PartNumberError(
            f"Buckle-partnummer '{scanned_buckle_part}' komt niet overeen met database voor workorder {workorder_id}."
        )

    # Geen return nodig: geen exception == alles OK


def write_trace_ids(workorder_id: str,
                    frame_trace: str | None,
                    belt_trace: str | None,
                    buckle_trace: str | None):

    wb, ws = _load_sheet()
    rows = _find_rows_for_workorder(ws, workorder_id)
    if not rows:
        raise ValueError(f"Geen rijen gevonden voor workorder '{workorder_id}' op pagina twee.")

    # Bestaande benchframe trace IDs controleren (uniek)
    existing_frame_traces = set()
    for r in range(2, ws.max_row + 1):
        part = str(ws[f"{COL_PARTID}{r}"].value or "").strip()
        trace = ws[f"{COL_TRACEID}{r}"].value
        if trace is not None:
            existing_frame_traces.add(str(trace).strip())
            print("trace in database is")

    # Frame trace mag niet al bestaan
    if frame_trace and frame_trace in existing_frame_traces:
        raise ValueError(f"Benchframe TRACEID '{frame_trace}' bestaat al; moet uniek zijn.")

    # Schrijven per rij
    for r in rows:
        desc = str(ws[f"{COL_DESCRIPTION}{r}"].value or "").lower()
        if frame_trace is not None and "benchframe" in desc:
            ws[f"{COL_TRACEID}{r}"] = frame_trace
        elif belt_trace is not None and "seatbelt" in desc:
            ws[f"{COL_TRACEID}{r}"] = belt_trace
        elif buckle_trace is not None and "buckle" in desc:
            ws[f"{COL_TRACEID}{r}"] = buckle_trace

    wb.save(WORKORDERS_PATH)
